import io
import os
import logging
from typing import BinaryIO

logger = logging.getLogger(__name__)

# Limit text extraction to ~1MB to prevent OOM on large files
MAX_TEXT_LENGTH = 1_000_000


def parse_pdf(file_obj: BinaryIO) -> str:
    try:
        import pypdf

        reader = pypdf.PdfReader(file_obj)
        output = io.StringIO()
        current_length = 0

        for page in reader.pages:
            text = page.extract_text() or ""
            if not text:
                continue

            output.write(text)
            output.write("\n")
            current_length += len(text) + 1

            if current_length >= MAX_TEXT_LENGTH:
                logger.warning(f"PDF extraction truncated at {MAX_TEXT_LENGTH} chars")
                break

        return output.getvalue()
    except ImportError:
        return "Error: pypdf no instalado"


def parse_docx(file_obj: BinaryIO) -> str:
    try:
        import docx

        doc = docx.Document(file_obj)
        output = io.StringIO()
        current_length = 0

        for p in doc.paragraphs:
            text = p.text
            if not text:
                continue

            output.write(text)
            output.write("\n")
            current_length += len(text) + 1

            if current_length >= MAX_TEXT_LENGTH:
                logger.warning(f"DOCX extraction truncated at {MAX_TEXT_LENGTH} chars")
                break

        return output.getvalue()
    except ImportError:
        return "Error: python-docx no instalado"


def parse_excel(file_obj: BinaryIO) -> str:
    try:
        import openpyxl

        if isinstance(file_obj, (bytes, bytearray)):
            file_obj = io.BytesIO(file_obj)

        # read_only=True is essential for performance and memory efficiency with large files
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        try:
            output = io.StringIO()
            current_length = 0

            # Iterating over worksheets directly is slightly more efficient than wb[sheet_name]
            for i, ws in enumerate(wb.worksheets):
                if i > 0:
                    output.write("\n")
                    current_length += 1

                # ws.iter_rows with values_only=True is the fastest way to get data
                for row in ws.iter_rows(values_only=True):
                    # Efficiently filter out empty rows and join values
                    if any(c is not None for c in row):
                        line = ",".join(str(c) if c is not None else "" for c in row)
                        output.write(line + "\n")
                        current_length += len(line) + 1

                        if current_length >= MAX_TEXT_LENGTH:
                            logger.warning(f"Excel extraction truncated at {MAX_TEXT_LENGTH} chars")
                            break

                if current_length >= MAX_TEXT_LENGTH:
                    break

            return output.getvalue()
        finally:
            wb.close()
    except ImportError:
        return "Error: openpyxl no instalado"


def parse_csv(file_obj: BinaryIO) -> str:
    if isinstance(file_obj, (bytes, bytearray)):
        file_obj = io.BytesIO(file_obj)

    # Use a generator to read lines and stop if we exceed the limit
    wrapper = io.TextIOWrapper(file_obj, encoding="utf-8", errors="ignore")
    output = io.StringIO()
    current_length = 0

    try:
        # Read line by line to keep memory usage low
        for line in wrapper:
            output.write(line)
            current_length += len(line)

            if current_length >= MAX_TEXT_LENGTH:
                logger.warning(f"CSV extraction truncated at {MAX_TEXT_LENGTH} chars")
                break

        return output.getvalue()
    finally:
        wrapper.detach()


def extract_text_content(file_obj: BinaryIO, filename: str) -> str:
    file_ext = os.path.splitext(filename)[1].lower()
    text_content = ""

    try:
        if file_ext == ".txt":
            wrapper = io.TextIOWrapper(file_obj, encoding="utf-8", errors="ignore")
            try:
                # Read with limit for .txt files as well
                text_content = wrapper.read(MAX_TEXT_LENGTH)
            finally:
                wrapper.detach()
        elif file_ext == ".pdf":
            text_content = parse_pdf(file_obj)
        elif file_ext == ".docx":
            text_content = parse_docx(file_obj)
        elif file_ext == ".xlsx":
            text_content = parse_excel(file_obj)
        elif file_ext == ".csv":
            text_content = parse_csv(file_obj)
        else:
            wrapper = io.TextIOWrapper(file_obj, encoding="utf-8", errors="ignore")
            try:
                text_content = wrapper.read(MAX_TEXT_LENGTH)
            finally:
                wrapper.detach()
    except Exception as exc:
        logger.exception("Failed to extract text from %s", filename)
        text_content = f"Binary content extracted placeholder (Error: {str(exc)})"

    if not text_content:
        text_content = "Empty content"

    return text_content
